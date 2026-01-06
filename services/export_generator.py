"""
Export Generator Service
Generates Excel reports for invoices and items
"""
import logging
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List

logger = logging.getLogger(__name__)


class ExportGenerator:
    """Service for generating Excel export reports."""
    
    def __init__(self):
        # Define colors
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_invoices_report(self, invoices: List) -> BytesIO:
        """
        Generate Excel report for invoices.
        
        Columns: # | المورد | الرقم الضريبي | رقم الفاتورة | التاريخ | 
                 المجموع الفرعي | الخصم | الضريبة | الإجمالي
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"
            ws.sheet_view.rightToLeft = True
            
            # Set column widths
            column_widths = [5, 30, 20, 15, 12, 15, 12, 12, 15]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:I{row}')
            ws[f'A{row}'] = f"تقرير الفواتير - {datetime.now().strftime('%Y-%m-%d')}"
            ws[f'A{row}'].font = self.title_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', readingOrder=2)
            row += 2
            
            # Headers
            headers = ["#", "المورد", "الرقم الضريبي", "رقم الفاتورة", "التاريخ", 
                      "المجموع الفرعي", "الخصم", "الضريبة", "الإجمالي"]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                cell.border = self.border
            
            row += 1
            
            # Data rows
            for idx, invoice in enumerate(invoices, start=1):
                values = [
                    idx,
                    invoice['supplier_name'],
                    invoice['tax_number'],
                    invoice['invoice_number'],
                    invoice['invoice_date'],
                    invoice['subtotal'],
                    invoice['discount'],
                    invoice['tax_amount'],
                    invoice['total_amount']
                ]
                
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                    cell.border = self.border
                
                row += 1
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            logger.info(f"Generated invoices report with {len(invoices)} records")
            return excel_file
            
        except Exception as e:
            logger.error(f"Failed to generate invoices report: {e}")
            raise
    
    def generate_items_report(self, items: List) -> BytesIO:
        """
        Generate Excel report for items.
        
        Columns: اسم الصنف | الكمية | الوحدة | سعر الوحدة | الإجمالي | التاريخ
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            ws.sheet_view.rightToLeft = True
            
            # Set column widths
            column_widths = [40, 12, 12, 15, 15, 15]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"تقرير الأصناف - {datetime.now().strftime('%Y-%m-%d')}"
            ws[f'A{row}'].font = self.title_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', readingOrder=2)
            row += 2
            
            # Headers
            headers = ["اسم الصنف", "الكمية", "الوحدة", "سعر الوحدة", "الإجمالي", "التاريخ"]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                cell.border = self.border
            
            row += 1
            
            # Data rows
            for item in items:
                values = [
                    item['item_name'],
                    item['quantity'],
                    item['unit'],
                    item['unit_price'],
                    item['total'],
                    item['invoice_date']
                ]
                
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                    cell.border = self.border
                
                row += 1
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            logger.info(f"Generated items report with {len(items)} records")
            return excel_file
            
        except Exception as e:
            logger.error(f"Failed to generate items report: {e}")
            raise


# Global instance
export_generator = ExportGenerator()

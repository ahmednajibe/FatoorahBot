"""
Excel Generator Service
Exports invoice data to Excel format
"""
import logging
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.invoice import InvoiceData

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Service for generating Excel files from invoice data."""
    
    def __init__(self):
        # Define colors
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self, invoice: InvoiceData) -> BytesIO:
        """
        Generate Excel file from invoice data.
        
        Returns:
            BytesIO: Excel file in memory
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice"
            # Set worksheet to RTL (Right-to-Left) for Arabic
            ws.sheet_view.rightToLeft = True
            
            # Set column widths
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            row = 1
            
            # Title
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "بيانات الفاتورة"
            ws[f'A{row}'].font = self.title_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', readingOrder=2)
            row += 2
            
            # Invoice header info
            info_data = [
                ("اسم المورد:", invoice.supplier_name or "غير محدد"),
                ("الرقم الضريبي:", invoice.tax_number or "غير محدد"),
                ("رقم الفاتورة:", invoice.invoice_number or "غير محدد"),
                ("التاريخ:", invoice.invoice_date or "غير محدد"),
            ]
            
            for label, value in info_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
            
            # Items table header
            headers = ["اسم الصنف", "الكمية", "الوحدة", "سعر الوحدة", "الإجمالي"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            row += 1
            
            # Items data
            for item in invoice.items:
                ws[f'A{row}'] = item.name
                ws[f'B{row}'] = item.quantity
                ws[f'C{row}'] = item.unit
                ws[f'D{row}'] = item.unit_price
                ws[f'E{row}'] = item.total
                
                # Center align all cells
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                    cell.border = self.border
                
                row += 1
            
            # Totals
            totals_data = [
                ("المجموع الفرعي:", invoice.subtotal),
                ("الخصم:", invoice.discount),
                ("الضريبة:", invoice.tax_amount),
                ("الإجمالي النهائي:", invoice.total_amount),
            ]
            
            for label, value in totals_data:
                ws[f'D{row}'] = label
                ws[f'D{row}'].font = Font(bold=True)
                ws[f'E{row}'] = value
                ws[f'E{row}'].font = Font(bold=True) if label.startswith("الإجمالي") else Font()
                row += 1
            
            # Validation message if present
            if invoice.validation_message:
                row += 1
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = f"التدقيق: {invoice.validation_message}"
                ws[f'A{row}'].font = Font(italic=True)
                ws[f'A{row}'].alignment = Alignment(horizontal='center', readingOrder=2)
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            logger.info(f"Excel file generated successfully for invoice {invoice.invoice_number}")
            return excel_file
            
        except Exception as e:
            logger.error(f"Failed to generate Excel: {e}")
            raise


# Global instance
excel_generator = ExcelGenerator()
